import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from io import BytesIO

# Common formatting functions
def add_heading(ws, text, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal="center")
    return row + 2

def write_df(ws, df, start_row):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    # Auto-width columns
    for col in ws.iter_cols(min_row=start_row, max_row=ws.max_row):
        max_length = max(len(str(cell.value)) for cell in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    return r_idx + 2

def generate_report(analysis_path, marks_path, params):
    """Unified report generator with parameterized configuration"""
    # Load and process data
    def load_data(file):
        df = pd.read_excel(file, header=[0, 1])
        df.columns = [f"{c[0]} ({c[1]})" if "Unnamed" not in c[0] else c[1] for c in df.columns]
        return df

    df_analysis = load_data(analysis_path)
    df_marks = load_data(marks_path)

    # 1) Subject-wise percentage table
    percent_dict = df_analysis.iloc[-1].to_dict()
    subject_data = []
    for i, (subj, col, col_type) in enumerate(params['subjects'], 1):
        row = {"S.N.": i, f"SUBJECT({params['semester']})": subj, 
               **{f"{t} %": "-" for t in ['TH', 'TW', 'PR', 'OR']}}
        if col in percent_dict and pd.notna(percent_dict[col]):
            row[f"{col_type} %"] = f"{percent_dict[col]:.2f}"
        subject_data.append(row)
    subject_df = pd.DataFrame(subject_data)

    # 2) Overall result summary
    total_students = int(df_analysis.iloc[0, -1])
    sgpa_data = {
        "ALL CLEAR": df_analysis.iloc[18, -1],
        "DISTINCTION (> 7.75 SGPA)": df_analysis.iloc[5, -1],
        "FIRST CLASS (6.75 TO 7.74 SGPA)": df_analysis.iloc[7, -1],
        "HIGH.SECOND CLASS (6.25 TO 6.74 SGPA)": df_analysis.iloc[9, -1],
        "SECOND CLASS (5.5 TO 6.24 SGPA)": df_analysis.iloc[11, -1],
        "PASS CLASS (4.0 TO 5.49 SGPA)": df_analysis.iloc[13, -1],
        "FAIL": df_analysis.iloc[15, -1]
    }
    overall_df = pd.DataFrame([(k, round(v, 2)) for k, v in sgpa_data.items()], 
                             columns=["RESULT", "NO OF STUDENTS (%)"])

    # 3) Class toppers
    df_marks.rename(columns={params['name_col']: "Name", params['sgpa_col']: "SGPA"}, inplace=True)
    df_marks["SGPA"] = pd.to_numeric(df_marks["SGPA"], errors='coerce').round(2)
    df_sorted = df_marks[["Name", "SGPA"]].dropna().sort_values("SGPA", ascending=False)
    
    top_5_sgpas = df_sorted["SGPA"].drop_duplicates().head(5)
    topper_rows = []
    for rank, sgpa in enumerate(top_5_sgpas, 1):
        for idx, name in enumerate(df_sorted[df_sorted["SGPA"] == sgpa]["Name"]):
            topper_rows.append([rank if idx == 0 else "", name, sgpa])
    topper_df = pd.DataFrame(topper_rows, columns=["Rank", "Name of Student", "SGPA"])

    # 4) Subject toppers
    subject_topper_data = []
    for sn, (col, subj) in enumerate(params['subject_map'].items(), 1):
        df_marks[col] = pd.to_numeric(df_marks[col], errors='coerce')
        if (max_marks := df_marks[col].max()) and not pd.isna(max_marks):
            students = df_marks[df_marks[col] == max_marks]["Name"].dropna().tolist()
            for idx, name in enumerate(students):
                subject_topper_data.append([
                    sn if idx == 0 else "", 
                    subj if idx == 0 else "", 
                    name, 
                    int(max_marks)
                ])
    subject_topper_df = pd.DataFrame(subject_topper_data, 
        columns=["S.N.", f"SUBJECT ({params['semester']})", "NAME OF THE STUDENT(S)", "MARKS OBTAINED (OUT OF 100)"])

    # Create Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = params['report_title']
    current_row = 1

    for section in [
        ("SUBJECT-WISE % RESULT", subject_df),
        ("OVERALL RESULT", overall_df, f"TOTAL NO OF STUDENTS APPEARED = {total_students}"),
        ("CLASS TOPPERS", topper_df),
        ("SUBJECT TOPPERS", subject_topper_df)
    ]:
        current_row = add_heading(ws, section[0], current_row)
        if section[0] == "OVERALL RESULT":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            ws.cell(row=current_row, column=1, value=section[2]).alignment = Alignment(horizontal="center")
            current_row += 2
        current_row = write_df(ws, section[1], current_row)

    return wb

# Semester configurations
CONFIG = {
    "SEM-III": {
        "subjects": [
            ("Discrete Mathematics", "DM (TH)", "TH"),
            ("Fundamentals of Data Structures", "FDS (TH)", "TH"),
            ("Object-Oriented Programming", "OOP (TH)", "TH"),
            ("Computer Graphics", "CG (TH)", "TH"),
            ("Digital Electronics and Logic Design", "DELD (TH)", "TH"),
            ("DSA TW", "DSA Lab (TW)", "TW"),
            ("DSA PR", "DSA Lab (PR)", "PR"),
            ("OOP & CG Lab TW", "OOPCG Lab (TW)", "TW"),
            ("OOP & CG Lab PR", "OOPCG Lab (PR)", "PR"),
            ("DEL TW", "DELD Lab (TW)", "TW"),
            ("BCS TW", "BCS (TW)", "TW"),
            ("HSS TW", "HSS (TW)", "TW")
        ],
        "name_col": "SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (Name)",
        "sgpa_col": "SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (SGPA)",
        "subject_map": {
            'SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (DM)': 'Discrete Mathematics',
            'SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (FDS)': 'Fundamentals of Data Structures',
            'SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (OOP)': 'Object Oriented Programming',
            'SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (CG)': 'Computer Graphics',
            'SE Comp Sem-I Result Analysis, Exam: Dec2023 (A.Y.2023-24) (DELD)': 'Digital Electronics & Logic Design'
        },
        "report_title": "SE_NOV_2023 Result Report",
        "semester": "SEM-III"
    },
    "SEM-IV": {
        "subjects": [
            ("Engineering Mathematics-III", "EM3 (TH)", "TH"),
            ("Data Structures and Algorithms", "DSA (TH)", "TH"),
            ("Software Engineering", "SE (TH)", "TH"),
            ("Microprocessors", "MP (TH)", "TH"),
            ("Principles of Programming Languages", "PPL (TH)", "TH"),
            ("EM3 Lab", "EM3 LAB (TW)", "TW"),
            ("DSA Lab TW", "DSA Lab (TW)", "TW"),
            ("DSA Lab PR", "DSA Lab (PR)", "PR"),
            ("MP Lab TW", "MP Lab (TW)", "TW"),
            ("MP Lab OR", "MP Lab (OR)", "OR"),
            ("Project-Based Learning 2", "PBL2 (TW)", "TW"),
            ("COC", "COC (TW)", "TW"),
        ],
        "name_col": "SE Comp Result Analysis Sem-II (Exam: May2024) (Name)",
        "sgpa_col": "SE Comp Result Analysis Sem-II (Exam: May2024) (SGPA)",
        "subject_map": {
            'SE Comp Result Analysis Sem-II (Exam: May2024) (EM3)': 'Engineering Mathematics-III',
            'SE Comp Result Analysis Sem-II (Exam: May2024) (DSA)': 'Data Structures & Algorithms',
            'SE Comp Result Analysis Sem-II (Exam: May2024) (SE)': 'Software Engineering',
            'SE Comp Result Analysis Sem-II (Exam: May2024) (MP)': 'Microprocessor',
            'SE Comp Result Analysis Sem-II (Exam: May2024) (PPL)': 'Principles of Programming Languages'
        },
        "report_title": "SE_MAY_2024 Result Report",
        "semester": "SEM-IV"
    },
    "SEM-V": {
        "subjects": [
            ("Database Management Systems", "DBMS (TH)", "TH"),
            ("Theory of Computation", "TOC (TH)", "TH"),
            ("Systems Programming and OS", "SPOS (TH)", "TH"),
            ("Cyber Security & Network Security", "CNS (TH)", "TH"),
            ("IoT & Embedded Systems (ELE I)", "IOT&ES (TH)", "TH"),
            ("Human Computer Interaction (ELE I)", "HCI (TH)", "TH"),
            ("Distributed Systems (ELE I)", "DS (TH)", "TH"),
            ("Seminar and TeleCom TW", "STC (TW)", "TW"),
            ("DBMS Lab TW", "DBMSLab (TW)", "TW"),
            ("DBMS Lab PR", "DBMSLab (PR)", "PR"),
            ("LP-I TW", "LP - I (TW)", "TW"),
            ("LP-I PR", "LP - I (PR)", "PR"),
            ("CNS Lab TW", "CNSLab (TW)", "TW"),
            ("CNS Lab OR", "CNSLab (OR)", "OR"),
        ],
        "name_col": "TE Result Analysis, Exam Nov-2023. (Name)",
        "sgpa_col": "TE Result Analysis, Exam Nov-2023. (SGPA)",
        "subject_map": {
            "TE Result Analysis, Exam Nov-2023. (DATABASE MANAGEMENT SYSTEMS)": "Database Management Systems",
            "TE Result Analysis, Exam Nov-2023. (THEORY OF COMPUTATION)": "Theory of Computation",
            "TE Result Analysis, Exam Nov-2023. (SYSTEM PROGRAMMING AND OS )": "Systems Programming and OS",
            "TE Result Analysis, Exam Nov-2023. (COMPUTER NETWORKS AND SECURITY)": "Computer Networks & Security",
            "TE Result Analysis, Exam Nov-2023. (INTERNET OF THINGS & EMBEDDED SYSTEMS)": "IoT & Embedded Systems",
            "TE Result Analysis, Exam Nov-2023. (HUMAN COMPUTER INTERFACE)": "Human Computer Interaction",
            "TE Result Analysis, Exam Nov-2023. (DISTRIBUTED SYSTEMS)": "Distributed Systems"
        },
        "report_title": "TE_NOV_2023 Result Report",
        "semester": "SEM-V"
    },
        "SEM-VI": {
        "subjects": [
            ("Data Science and Big Data Analytics", "DSBDA (TH)", "TH"),
            ("Web Technologies", "WT (TH)", "TH"),
            ("Artificial Intelligence", "AI (TH)", "TH"),
            ("Cloud Computing (ELE II)", "CC (TH)", "TH"),
            ("Information Security (ELE II)", "IS (TH)", "TH"),
            ("Augmented & Virtual Reality (ELE II)", "A&VR (TH)", "TH"),
            ("Internship", "INTERNSHIP (TW)", "TW"),
            ("DSBDA Lab", "DSBDA Lab (TW)", "TW"),
            ("DSBDA Lab", "DSBDA Lab (PR)", "PR"),
            ("LP-II Lab", "LP-II LAB (TW)", "TW"),
            ("LP-II Lab", "LP-II Lab (PR)", "PR"),
            ("Web Technologies Lab", "WT Lab (TW)", "TW"),
            ("Web Technologies Lab", "WT Lab (OR)", "OR"),
        ],
        "name_col": "TE Comp Result Analysis Sem-II (Exam: May2024) (Name)",
        "sgpa_col": "TE Comp Result Analysis Sem-II (Exam: May2024) (SGPA)",
        "subject_map": {
            'TE Comp Result Analysis Sem-II (Exam: May2024) (DSBDA)': 'Data Science & Big Data Analytics',
            'TE Comp Result Analysis Sem-II (Exam: May2024) (WT)': 'Web Technology',
            'TE Comp Result Analysis Sem-II (Exam: May2024) (AI)': 'Artificial Intelligence',
            'TE Comp Result Analysis Sem-II (Exam: May2024) (CC)': 'Cloud Computing (ELE II)',
            'TE Comp Result Analysis Sem-II (Exam: May2024) (IS)': 'Information Security (ELE II)',
            'TE Comp Result Analysis Sem-II (Exam: May2024) (A&VR)': 'Augmented & Virtual Reality (ELE II)'
        },
        "report_title": "TE_MAY_2024 Result Report",
        "semester": "SEM-VI"
    },

    "SEM-VII": {
        "subjects": [
    ("Design & Analysis Of Algorithms (DAA)", "DAA (TH)", "TH"),
    ("Machine Learning (ML)", "ML (TH)", "TH"),
    ("Blockchain Technology (BT)", "BT (TH)", "TH"),
    ("EL-III : Pervasive Computing (PC)", "PC (ELE-III) (TH)", "TH"),
    ("EL-III : Multimedia Techniques (MT)", "MT (ELE-III) (TH)", "TH"),
    ("EL-III: Cyber Security and Digital Forensics (CSDF) ", "CSDF (ELE-III) (TH)", "TH"),
    ("EL-III: Object Oriented Modelling & Design (OOMD)", "OOMD (ELE-III) (TH)", "TH"),
    ("EL-IV: Information Retrieval (IR)", "IR (ELE-IV) (TH)", "TH"),
    ("EL-IV: Mobile Computing (MC)", "MC (ELE-IV) (TH)", "TH"),
    ("EL-IV: Software Testing & Quality Assurance (STQA)", "STQA (ELE-IV) (TH)", "TH"),
    ("LP-III (TW)", "LP-III (TW)", "TW"),
    ("LP-III (PR)", "LP-III (PR)", "PR"),
    ("LP-IV (TW)", "LP-IV (TW)", "TW"),
    ("Project-I (TW)", "PROJECT STAGE - I (TW)", "TW"),
        ],
        "name_col": "BE COMP Result Analysis - Exam Nov-2023 (Name)",
        "sgpa_col": "BE COMP Result Analysis - Exam Nov-2023 (SGPA)",
        "subject_map": {
            "BE COMP Result Analysis - Exam Nov-2023 (DESIGN & ANALYSIS OF ALGO)": "Design & Analysis of Algorithm",
            "BE COMP Result Analysis - Exam Nov-2023 (MACHINE LEARNING)": "Machine Learning",
            "BE COMP Result Analysis - Exam Nov-2023 (BLOCKCHAIN TECHNOLOGY)": "Blockchain Technology",
            "BE COMP Result Analysis - Exam Nov-2023 (PERVASIVE COMPUTING)": "Pervasive Computing",
            "BE COMP Result Analysis - Exam Nov-2023 (MULTIMEDIA TECHNIQUES)": "Multimedia Techniques",
            "BE COMP Result Analysis - Exam Nov-2023 (CYBER SEC & DIGITAL FORENSICS)": "Cyber Security & Digital Forensics",
            "BE COMP Result Analysis - Exam Nov-2023 (OBJ. ORIENTED MODL. & DESIGN)": "Object Oriented Modeling & Design",
            "BE COMP Result Analysis - Exam Nov-2023 (INFORMATION RETRIEVAL)": "Information Retrieval",
            "BE COMP Result Analysis - Exam Nov-2023 (MOBILE COMPUTING)": "Mobile Computing",
            "BE COMP Result Analysis - Exam Nov-2023 (SOFTWARE TESTING & QUALITY ASSURANCE)": "Software Testing & QA"
        },
        "report_title": "BE_NOV_2023 Result Report",
        "semester": "SEM-VII"
    },

    "SEM-VIII": {
        "subjects": [
    ("High Performance Computing", "HPC (TH)", "TH"),
    ("Deep Learning", "DL (TH)", "TH"),
    ("Natural Language Processing (ELE V)", "NLP (TH)", "TH"),
    ("Image Processing (ELE V)", "IP (TH)", "TH"),
    ("Pattern Recognition (ELE VI)", "PR (TH)", "TH"),
    ("Business Intelligence (ELE VI)", "BI (TH)", "TH"),
    ("LP-V Lab (TW)", "LP-V (TW)", "TW"),
    ("LP-V Lab (PR)", "LP-V (PR)", "PR"),
    ("LP-VI Lab (TW)", "LP-VI (TW)", "TW"),
    ("Project Stage II (TW)", "PROJECT-II (TW)", "TW"),
    ("Project Stage II (OR)", "PROJECT-II (OR)", "OR")
        ],
        "name_col": "Result Analysis BE Comp May-2024 (Name)",
        "sgpa_col": "Result Analysis BE Comp May-2024 (BE SGPA)",
        "subject_map": {
            'Result Analysis BE Comp May-2024 (HPC)': 'High Performance Computing',
            'Result Analysis BE Comp May-2024 (DL)': 'Deep Learning',
            'Result Analysis BE Comp May-2024 (NLP)': 'Natural Language Processing',
            'Result Analysis BE Comp May-2024 (IP)': 'Image Processing',
            'Result Analysis BE Comp May-2024 (PR)': 'Pattern Recognition',
            'Result Analysis BE Comp May-2024 (BI)': 'Business Intelligence'
        },
        "report_title": "BE_MAY_2024 Result Report",
        "semester": "SEM-VIII"
    }
}

# Streamlit UI
st.title("NAAC Report Generator 📊")
st.header("1. Upload Files")
analysis_file = st.file_uploader("Analysis Excel", type=["xlsx"])
marks_file = st.file_uploader("Result Excel", type=["xlsx"])
semester = st.selectbox("2. Select Semester", ["SEM-III", "SEM-IV", "SEM-V", "SEM-VI", "SEM-VII", "SEM-VIII"])
# In the Streamlit UI section where you create the download button:
# Enhanced filename generation in download button
if st.button("Generate Report") and analysis_file and marks_file:
    with st.spinner("Generating..."):
        try:
            config = CONFIG[semester]
            wb = generate_report(analysis_file, marks_file, config)
            
            # Generate filename with proper prefix and date
            prefix = config['report_title'].split('_')[0]  # SE/TE/BE
            date_code = 'NOV23' if 'NOV' in config['report_title'] else 'May24'
            
            output = BytesIO()
            wb.save(output)
            
            st.success("Done!").download_button(
                "Download Report",
                output.getvalue(),
                f"{prefix}_NAAC_{date_code}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error generating report: {str(e)}")

st.markdown("### Instructions\n1. Upload both files\n2. Select semester\n3. Generate and download")